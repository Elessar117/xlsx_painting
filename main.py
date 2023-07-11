import openpyxl
from openpyxl.styles import PatternFill

dark_green_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
light_green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
dark_red_fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid')
light_red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

workbook = openpyxl.load_workbook('screening_table_spot.xlsx')

worksheet = workbook.active

columns_to_compare = {'D': 'L', 'E': 'M', 'F': 'N', 'G': 'O', 'H': 'P', 'I': 'Q', 'J': 'R', 'K': 'S'}

total_comparisons = 0
correct_comparisons = 0

row = 2
while worksheet['D' + str(row)].value is not None:
    for col, counterpart in columns_to_compare.items():
        if worksheet[counterpart + str(row)].value is not None:
            total_comparisons += 1
            if worksheet[col + str(row)].value == worksheet[counterpart + str(row)].value:
                worksheet[col + str(row)].fill = dark_green_fill
                worksheet[counterpart + str(row)].fill = light_green_fill
                correct_comparisons += 1
            else:
                worksheet[col + str(row)].fill = dark_red_fill
                worksheet[counterpart + str(row)].fill = light_red_fill
    row += 1

percentage_correct = (correct_comparisons / total_comparisons) * 100

worksheet['W1'] = "Result"
worksheet['W1'] = "{}%".format(percentage_correct)

workbook.save('screening_table_spot_new.xlsx')
